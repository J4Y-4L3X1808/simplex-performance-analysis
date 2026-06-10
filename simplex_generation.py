import numpy as np
import timeit
import openpyxl
try: 
    from tqdm.auto import tqdm
except ImportError:
    class tqdm_t(object):
        def update(self, u):
            pass
    def tqdm(**kwargs):
        return tqdm_t()

def simplex(tableau, basis, pivot_rule):
    """
    Perform the single phase simplex method, assuming the tableau is in basic form.
    
    Parameters
    ----------
    tableau : array
        The simplex method tableau in basic form
    basis : array
        The list of indexes (in order) forming the basic set
    pivot_rule : function
        Given a tableau, returns the pivot index pair
    
    Returns
    -------
    soln : dict
        The solution dictionary with result, objective value, and solution.
    """
    M, N = tableau.shape
    m = M-1  # Number of constraints
    n = N-1  # Number of variables
    soln = {
        'result' : 'infeasible',
        'objective' : np.inf,
        'solution' : np.zeros((n,)),
        'basis' : basis
    }
    while np.any(tableau[-1, :-1] > 0):
        r, s = pivot_rule(tableau)
        if r == -1:  # No suitable pivot: unbounded
            soln['result'] = 'unbounded'
            return soln
        # Pivot
        tableau[r, :] = tableau[r, :] / tableau[r, s]
        for row in range(m+1):
            if r != row:
                tableau[row, :] = tableau[row, :] - tableau[row, s] * tableau[r, :]
        basis[r] = s
    # All reduced costs non-negative: optimal solution found
    soln['result'] = 'optimal'
    soln['objective'] = -tableau[-1, -1]
    soln['solution'] = tableau[:-1, -1]
    soln['basis'] = basis
    return soln

def setup(m, n):
    """
    Create a tableau in basic form of size (m, n).

    Parameters
    ----------
    m : integer
        Number of constraints (rows in the tableau). 
    n : integer
        Number of decision variables (columns in tableau). 

    Returns
    -------
    tableau : array
        The (m+1)x(n+1) tableau in basic form.
    basis : array
        The basis (always [n, ..., n+m]).
    """
    tableau = np.zeros((m+1, n+m+1), dtype=np.float64)
    tableau[:-1, -1] = np.random.rand(m)
    tableau[-1, :n] = np.random.rand(n)
    tableau[:-1, :n] = np.random.rand(m, n)
    tableau[:-1, n:-1] = np.identity(m)
    basis = np.array(range(n, n+m))
    return tableau, basis

def run_times(student_id, find_pivot_bland, find_pivot_largest_subscript, find_pivot_largest_change, time_largest_subscript=True, time_largest_change=True):
    """
    Create an Excel spreadsheet containing the runtimes of the Simplex Method.

    Parameters
    ----------
    student_id : int
        The student ID used as a random seed
    find_pivot_bland : function
        Function implementing Bland's rule for pivoting
    find_pivot_largest_subscript : function
        Function implementing the largest coefficient rule for pivoting
    find_pivot_largest_change : function
        Function implementing the largest change rule for pivoting
    time_largest_subscript : boolean
        If true then produce timing data for Largest Subscript rule
    time_largest_change : boolean
        If true then produce timing data for Largest Change rule

    Returns
    -------
    success : boolean
        Did it work.
    """
    wb = openpyxl.Workbook()
    ws_ss = wb.active
    ws_ss.title = "Bland"
    wss = [ws_ss]
    if time_largest_subscript:
        ws_ls = wb.create_sheet("LargestSubscript")
        wss.append(ws_ls)
    if time_largest_change:
        ws_sc = wb.create_sheet("LargestChange")
        wss.append(ws_sc)
    for ws in wss:
        ws.cell(row=1, column=1, value="n")
    ns = np.array([ 12,  15,  18,  22,  26,  31,  38,  46,  55,  66,  79,  95, 114, 137, 164, 197, 237, 284, 341])
    n_instances = 50
    for ws in wss:
        for j in range(1, n_instances+1):
            ws.cell(row=1, column=j+1, value=f"Instance{j}")
    t = tqdm(total=int(np.sum(ns**1.5)))
    for i, n in enumerate(ns):
        for ws in wss:
            ws.cell(row=i+2, column=1, value=n)
        for j in range(n_instances):
            np.random.seed(j+student_id)
            time_ss = timeit.timeit("simplex(tableau, basis, pivot_rule=find_pivot_bland)",
                                     setup=f"tableau, basis = setup({n+5},{n})",
                                     globals={
                                         'simplex' : simplex,
                                         'find_pivot_bland' : find_pivot_bland,
                                         'setup'   : setup,
                                         'n'       : n,
                                             },
                                     number=10)
            ws_ss.cell(row=i+2, column=j+2, value=time_ss)
            if time_largest_subscript:
                np.random.seed(j+student_id)
                time_ls = timeit.timeit("simplex(tableau, basis, pivot_rule=find_pivot_largest_subscript)",
                                         setup=f"tableau, basis = setup({n+5},{n})",
                                         globals={
                                         'simplex' : simplex,
                                         'find_pivot_largest_subscript' : find_pivot_largest_subscript,
                                         'setup'   : setup,
                                         'n'       : n,
                                             },
                                         number=10)
                ws_ls.cell(row=i+2, column=j+2, value=time_ls)
            if time_largest_change:
                np.random.seed(j+student_id)
                time_sc = timeit.timeit("simplex(tableau, basis, pivot_rule=find_pivot_largest_change)",
                                         setup=f"tableau, basis = setup({n+5},{n})",
                                         globals={
                                         'simplex' : simplex,
                                         'find_pivot_largest_change' : find_pivot_largest_change,
                                         'setup'   : setup,
                                         'n'       : n,
                                             },
                                         number=10)
                ws_sc.cell(row=i+2, column=j+2, value=time_sc)
        t.update(int(n**1.5))
    success = wb.save('timings.xlsx')
    return success